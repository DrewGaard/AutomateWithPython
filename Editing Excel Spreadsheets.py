import openpyxl

wb = openpyxl.Workbook()

print(wb.sheetnames)

sheet = wb['Sheet']

print(sheet)

sheet['A1'] = 42

wb.create_sheet()

names = wb.sheetnames

print(names)

sheet2 = names[1]

print(sheet2)

wb[sheet2].title = "My New Sheet Name"

wb.save('exampleWB.xlsx')
